"""utils/output_xlsx.py — xlsx 清单输出

从 output_format.py 按介质拆分而来（xlsx 部分）。公开函数 export_to_one_excel_with_format
（openpyxl 引擎，支持向已有文件追加 sheet）。经 output_format.py 聚合入口导入即可。
"""
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def _str_width(s):
    """估算字符串在 Excel 中的显示宽度。非 ASCII 字符计 2 单位，ASCII 计 1。"""
    w = 0.0
    for ch in str(s):
        w += 2.0 if ord(ch) > 0x7F else 1.0
    return w


def _apply_col_widths(ws, col_widths):
    """将预计算的列宽写入 worksheet（openpyxl）。"""
    for c, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(c + 1)].width = width


def _sanitize_sheet_name(name: str) -> str:
    """清理 Excel sheet name 中的全角冒号（Excel 不支持 `：`）。"""
    return name.replace("：", "-")


# ── export_to_one_excel_with_format (openpyxl) ────────────────────────


def export_to_one_excel_with_format(df, output_path, sheet_name, title_name=None, add_title=True):
    """向指定文件写入一个 sheet（已存在则覆盖，否则新建），使用 openpyxl。"""
    sheet_name = _sanitize_sheet_name(sheet_name)
    if os.path.exists(output_path):
        workbook = load_workbook(output_path)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        worksheet = workbook.create_sheet(sheet_name)
    else:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    num_cols = df.shape[1]
    rows, cols = df.shape
    text_cols = {i for i, c in enumerate(df.columns) if df[c].dtype == object}
    col_widths = [_str_width(c) for c in df.columns]

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center", vertical="center")
    data_font = Font(size=10)
    data_alignment = Alignment(vertical="center")
    data_border = Border(
        left=Side(border_style="thin"), right=Side(border_style="thin"),
        top=Side(border_style="thin"), bottom=Side(border_style="thin"),
    )

    header_row = 2 if add_title else 1
    if add_title and title_name:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        worksheet["A1"].value = title_name
        worksheet["A1"].font = title_font
        worksheet["A1"].alignment = title_alignment

    for c, col_name in enumerate(df.columns):
        cell = worksheet.cell(row=header_row, column=c + 1)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    data_start = header_row + 1
    for r in range(rows):
        for c in range(cols):
            val = df.iloc[r, c]
            if pd.isna(val):
                val = ""
            else:
                col_widths[c] = max(col_widths[c], _str_width(val))
            cell = worksheet.cell(row=data_start + r, column=c + 1)
            cell.value = val
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
            if c in text_cols:
                cell.number_format = '@'

    _apply_col_widths(worksheet, [min(w + 2, 60) for w in col_widths])

    if rows >= 1 and cols >= 1:
        last_col = get_column_letter(cols)
        worksheet.auto_filter.ref = f"A{header_row}:{last_col}{data_start + rows - 1}"

    workbook.save(output_path)
    print(f"Sheet '{sheet_name}' 已成功导出至: {output_path}")
