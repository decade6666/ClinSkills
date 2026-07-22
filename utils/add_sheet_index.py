from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

def add_sheet_index(excel_path, index_sheet_name="目录"):
    """
    为已有 Excel 文件添加目录页（放在最前面），不改变原有 sheet 样式。
    目录包含两列：序号、表单名称（带跳转链接），并加上细框线。
    """

    wb = load_workbook(excel_path)

    # 如果目录页已存在，先删除再重建
    if index_sheet_name in wb.sheetnames:
        del wb[index_sheet_name]

    # 创建目录页并插入到最前
    index_ws = wb.create_sheet(index_sheet_name, 0)

    # 获取所有原有 sheet 名（不包括目录页）
    sheet_names = wb.sheetnames[1:]

    # 写入表头
    index_ws["A1"] = "序号"
    index_ws["B1"] = "表单名称"

    # 表头加粗 
    index_ws["A1"].font = Font(bold=True) 
    index_ws["B1"].font = Font(bold=True)

    # 写入目录内容
    for i, name in enumerate(sheet_names, start=1):
        row = i + 1
        index_ws[f"A{row}"] = i
        cell = index_ws[f"B{row}"]
        cell.value = name
        cell.hyperlink = f"#{name}!A1"   # 跳转到对应 sheet
        cell.style = "Hyperlink"

    # 自动列宽
    index_ws.column_dimensions["A"].width = 8
    index_ws.column_dimensions["B"].width = 30

    # 加上细框线
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    max_row = len(sheet_names) + 1
    for row in index_ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = border

    # 冻结首行 
    index_ws.freeze_panes = "A2"

    # 开启筛选 
    index_ws.auto_filter.ref = f"A1:B{max_row}"
    
    # 保存
    wb.save(excel_path)
