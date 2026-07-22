import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def summarize_excel(filepath: str, summary_title: str = "汇总表",
                    name_header: str = "核查表名称",
                    count_header: str = "质疑条数") -> None:
    """
    读取 Excel 文件，统计每个 sheet 的数据行数（不含标题行），
    在第一个位置插入汇总 sheet，并删除空数据 sheet。
    原文件所有样式信息保持不变。

    Args:
        filepath: Excel 文件路径（原地修改）
        summary_title: 汇总 sheet 名称，默认"汇总表"
        name_header: 名称列表头，默认"核查表名称"
        count_header: 计数列表头，默认"质疑条数"（DMR 语境；其他场景按需传入）
    """
    wb = openpyxl.load_workbook(filepath)

    original_sheets = wb.sheetnames  # 记录原有 sheet 名称列表

    # 统计每个 sheet 的数据行数（除第一行标题外）
    summary_data = []
    sheets_to_delete = []

    for name in original_sheets:
        ws = wb[name]
        # 检查"数据行"是否真的有内容（排除全空行被 openpyxl 计入的情况）
        actual_data_rows = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if any(cell.value is not None for cell in row):
                actual_data_rows += 1

        summary_data.append((name, actual_data_rows))
        if actual_data_rows == 0:
            sheets_to_delete.append(name)

    # 创建汇总 sheet，插入到最前面
    summary_ws = wb.create_sheet(title=summary_title, index=0)

    # 写入标题行
    summary_ws["A1"] = name_header
    summary_ws["B1"] = count_header

    # 标题行样式
    header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    white_font = Font(bold=True, name="Arial", size=11, color="FFFFFF")

    for cell in [summary_ws["A1"], summary_ws["B1"]]:
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据行
    data_font = Font(name="Arial", size=11)
    data_alignment_left = Alignment(horizontal="left", vertical="center")
    data_alignment_center = Alignment(horizontal="center", vertical="center")

    for i, (sheet_name, row_count) in enumerate(summary_data, start=2):
        summary_ws.cell(row=i, column=1, value=sheet_name).font = data_font
        summary_ws.cell(row=i, column=1).alignment = data_alignment_left
        summary_ws.cell(row=i, column=1).border = thin_border

        summary_ws.cell(row=i, column=2, value=row_count).font = data_font
        summary_ws.cell(row=i, column=2).alignment = data_alignment_center
        summary_ws.cell(row=i, column=2).border = thin_border

    # 自动列宽
    summary_ws.column_dimensions["A"].width = 30
    summary_ws.column_dimensions["B"].width = 15

    # 删除空数据 sheet
    for name in sheets_to_delete:
        del wb[name]

    wb.save(filepath)
    print("完成！汇总表已插入为第一个 sheet。")
    print(f"删除了 {len(sheets_to_delete)} 个空 sheet：{sheets_to_delete}")
    print("各 sheet 统计结果：")
    for name, count in summary_data:
        status = "（已删除）" if name in sheets_to_delete else ""
        print(f"  {name}: {count} 条{status}")


# 使用示例
if __name__ == "__main__":
    summarize_excel("your_file.xlsx")