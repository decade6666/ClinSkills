"""
build-metadata.py
EDC 元数据 Excel 解析入口。根据 edcType 调用对应的解析模块，生成 JSON 文件。

Usage: python build-metadata.py <edcType> <excelPath>
  edcType:  taimei5 | cmis | taimei6 | clinflash
  excelPath: 元数据 Excel 文件路径

JSON 输出在 Excel 同目录下，文件名由各解析模块决定。
"""
import sys
import os

try:
    import openpyxl
except ImportError:
    print("错误: 缺少 openpyxl，请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# 同目录下的解析模块
from parse_taimei5 import parse as parse_taimei5
from parse_cmis import parse as parse_cmis
from parse_taimei6 import parse as parse_taimei6
from parse_clinflash import parse as parse_clinflash
from _compat import load_workbook_patched

PARSERS = {
    "taimei5": parse_taimei5,
    "cmis": parse_cmis,
    "taimei6": parse_taimei6,
    "clinflash": parse_clinflash,
}


def main():
    if len(sys.argv) < 3:
        print("Usage: python build-metadata.py <edcType> <excelPath>", file=sys.stderr)
        print(f"  edcType: {', '.join(PARSERS.keys())}", file=sys.stderr)
        sys.exit(1)

    edc_type = sys.argv[1]
    excel_path = os.path.abspath(sys.argv[2])

    if edc_type not in PARSERS:
        print(f"无效的 EDC 类型: {edc_type}，支持: {', '.join(PARSERS.keys())}", file=sys.stderr)
        sys.exit(1)

    if not os.path.isfile(excel_path):
        print(f"文件不存在: {excel_path}", file=sys.stderr)
        sys.exit(1)

    print(f"读取元数据: {excel_path}")
    wb = load_workbook_patched(excel_path)
    print(f"  Sheets: {', '.join(wb.sheetnames)}")

    # 各解析模块返回 dict[str, dict]，key 为输出文件名（不含 .json），value 为数据
    output_dir = os.path.dirname(excel_path)
    results = PARSERS[edc_type](wb)
    wb.close()

    if not results:
        print("警告: 解析结果为空", file=sys.stderr)
        sys.exit(1)

    # 写入 JSON 文件
    import json
    from datetime import datetime, timezone

    for filename, data in results.items():
        out_path = os.path.join(output_dir, f"{filename}.json")
        data["_meta"] = {
            "edcType": edc_type,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceFile": excel_path,
        }
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        # 记录数 = 主 wrapper 键（非 _meta）的条目数，呼应 SKILL Step 4「各 section 记录数」
        n_records = sum(len(v) for k, v in data.items() if k != "_meta")
        print(f"  → {out_path}  ({n_records} 条)")


if __name__ == "__main__":
    main()
