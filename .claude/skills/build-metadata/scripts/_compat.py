"""
_compat.py
解析器共用工具：openpyxl 兼容 patch、通用读取函数、hasOther 判断。
"""
import openpyxl


# ── openpyxl 兼容 patch ────────────────────────────────────────

_patched = False


def ensure_openpyxl_patch():
    """patch openpyxl MatchPattern，兼容含非法 AutoFilter XML 的 Excel。

    太美6 的 Excel 已知存在此问题，其他 EDC 类型作为兜底也加载此 patch。
    """
    global _patched
    if _patched:
        return
    import openpyxl.descriptors.base as base_mod
    original_set = base_mod.MatchPattern.__set__

    def lenient_set(self, instance, value):
        try:
            original_set(self, instance, value)
        except ValueError:
            instance.__dict__[self.name] = value

    base_mod.MatchPattern.__set__ = lenient_set
    _patched = True


def load_workbook_patched(path):
    """加载 Excel，自动应用 openpyxl 兼容 patch。"""
    ensure_openpyxl_patch()
    return openpyxl.load_workbook(path, data_only=True)


# ── 通用 sheet 读取 ────────────────────────────────────────────


def read_sheet(wb, sheet_name):
    """读取 sheet 为 dict 列表，跳过全空行。"""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"_col{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        result.append(record)
    return result


# ── hasOther 判断 ──────────────────────────────────────────────


def has_other(display_value):
    """判断编码表枚举值是否为"其他"选项。

    匹配规则：
    - displayValue 以"其他"或"其它"开头（中文）
    - displayValue 等于 "other"（英文，不区分大小写）
    """
    dv = str(display_value) if display_value else ""
    return dv.startswith("其他") or dv.startswith("其它") or dv.lower() == "other"
